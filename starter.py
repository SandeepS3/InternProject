# Sandeep Singh

# Libraries Needed
# import pandas as pd
# import re
import openpyxl
import time
import random

# namelst = []
# #dataframe1 = pd.read_excel('test.xlsx')
# dataframe = openpyxl.load_workbook("test.xlsx")

# dataframe1 = dataframe.active
# #names into list
# for name in range(2, dataframe1.max_row+1):
#     names = dataframe1.cell(name,2)
#     namelst.append(names.value)

# for j in range(2, dataframe1.max_row+1):
#     answer = dataframe1.cell(j,4)
#     nanswer = re.sub(",", "", str(answer.value))
#     if "Stow" in nanswer.split():
#         print(namelst[j-2])
#         print(nanswer.split())
# print(namelst)

# def gui:
# from tkinter import *
# canvas = Tk()

# canvas.geometry("500x250")

# label = Label(canvas, text = "How many roles are we assigning?")
# label.pack()

# numroles = Entry(canvas, width = 50)
# numroles.pack()

# button1 = Button(canvas)

# canvas.mainloop()


# Start

# Currently 19 roles total
lst = [
    "RSRTD",
    "RSRPS",
    "RSRWS",
    "RSROP",
    "Decanters",
    "DecantPS",
    "DecantWS",
    "DockRun",
    "DockCRun",
    "DockLL",
    "DockCPB",
    "DockDS",
    "DockPD",
    "DockLI",
    "DockIDRT",
    "DockUPPC",
    "DockPS",
    "DockCrets",
    "Clerk",
]
priolst = [
    "Clerk",
    "RSRTD",
    "RSROP",
    "DecantPS",
    "DockPD",
    "DockIDRT",
    "DockPS",
    "DoctCrets",
]


# Might still have to ask which day of the week it is
def startup():
    numofroles = input("How many roles will be used today? ")
    roles = {}

    # Running Through which roles are needed and how many for them
    for num in range(int(numofroles)):
        rolename = input("What is the role name? ")
        while True:
            if rolename in lst:
                break
            print("Invalid input, please try again!")
            time.sleep(1)
            rolename = input("What is the role name? ")
        pplrole = input("How many are needed for " + rolename + " today? ")
        roles[rolename] = pplrole
    return roles


# Reading Excel

# Reading the excel to create a dictionary


def retrieveperson(roles):
    def personlst():
        aalst = []

        # Going through the priority
        for rl in range(2, spdsheet.max_row + 1):
            rls = spdsheet.cell(rl, 4)
            if priolstloop == rls.value:
                aalst.append(rl)
            # while (roles(priolstloop) != 0):
        return aalst

    ss = openpyxl.load_workbook("test.xlsx")
    spdsheet = ss.active

    # First loop through all the priority roles
    for priolstloop in priolst:
        # Check if this role is needed for the day, if it is continue on
        if priolstloop in roles:
            ppl = personlst()
            while roles[priolstloop] != 0:
                if len(ppl) == 0:
                    print(
                        "Not enough people for: "
                        + priolstloop
                        + ", still need "
                        + str(roles[priolstloop])
                        + " more people!"
                    )
                    break

                person = random.choice(ppl)
                spdsheet.cell(person, 8).value = priolstloop
                roles[priolstloop] = int(roles[priolstloop]) - 1
                ppl.remove(person)

            print(roles)
            ss.save("test.xlsx")


# Make a way to ensure people with more then one prio can get accesed (use str modifiers)
# Make a way so that people already assigned do not get reassigned (a list of total people)


r1 = startup()
retrieveperson(r1)
