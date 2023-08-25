# Sandeep Singh

# Libraries Needed
import re
import openpyxl
import time
import random

#Lists that contain roles
lst = [
    "RSRTD",
    "RSRPS",
    "RSRWS",
    "RSROP",
    "Decanter",
    "DecantPS",
    "DecantWS",
    "TrailerUL",
    "DockRun",
    "DockCRun",
    "DockLL",
    "DockCPB",
    "DockDS",
    "DockPD",
    "DockLI",
    "DockIDRT",
    "DockUPPC",
    "DockPS",
    "DockPG",
    "DockDeS",
    "Crets",
    "Clerk",
    "Stower",
    "PalletD",
]
priolst = [
    "Clerk",
    "DockPG",
    "DockDeS",
    "RSRTD",
    "RSROP",
    "RSRPS",
    "DecantPS",
    "DockPD",
    "DockIDRT",
    "DockPS",
    "Crets",
    "Stower",
    "PalletD",
]
usedppl = []

#Put roles and amount of people needed into dictionary
def startup():
    roles = {}
    for row in range(2, spdsheet2.max_row + 1):
        rolename = (spdsheet2.cell(row, 1).value)
        pplrole = (spdsheet2.cell(row, 2).value)
        if str(pplrole) != "None":
            if int(pplrole) > 0: roles[rolename] = pplrole


    return roles

# Reading Excel
# Reading the excel to create a dictionary

#Start of Shift
def retrievepersonb4(roles):
    def personlst():
        #oldregaalst not used anymore, only kept for less code editing
        prioaalst = []
        oldregaalst = []
        decanthourlst = []
        decantprio = []
        trailerprio = []
        unusedppl = []

        # Going through the priority
        for row in range(2, spdsheet.max_row + 1):
            rls = spdsheet.cell(row, 4)
            rls1 = re.sub(",", "", str(rls.value))
            # Checks for all their prios and makes sure person not used already
            if (priolstloop in rls1.split()) and (row not in usedppl):
                prioaalst.append(row)

            if ("TrailerUL" in rls1.split()) and (row not in usedppl):
                trailerprio.append(row)

            if ("Decanter" in rls1.split()) and (row not in usedppl):
                decantprio.append(row)

            dcnt = spdsheet.cell(row, 5).value
            if (str(dcnt) != "None"):
                if (int(dcnt) < 9) and (row not in usedppl):
                    decanthourlst.append(row)
            
            if (row not in usedppl):
                unusedppl.append(row)

        return prioaalst, oldregaalst, decanthourlst, decantprio, unusedppl, trailerprio

    #Cancel out people who are not apart of the shift for the day
    clockedinppl = []
    for row in range(2, spdsheet3.max_row + 1):
        rls3 = spdsheet3.cell(row, 2)
        if str(rls3.value) != "None": clockedinppl.append(rls3.value)

    for row in range(2, spdsheet.max_row + 1):
        rls2 = spdsheet.cell(row, 9)
        rls3 = spdsheet.cell(row, 3)
    
        if str(rls2.value) == "X" or str(rls2.value) == "x": 
            if row not in usedppl: usedppl.append(row)
        if str(rls3.value) not in clockedinppl:
            if not (str(rls2.value) == "X" or str(rls2.value) == "x"):
                spdsheet.cell(row, 7).value = "Clocked in After, See Manager!!"
                if row not in usedppl: usedppl.append(row)
    
    #Make decant its own variable
    d = "Decanter"
    if d in roles: 
        decant = roles[d]
        roles.pop(d)

    # First loop through all the priority roles
    for priolstloop in priolst:
        # Check if this role is needed for the day, if it is continue on
        if priolstloop in roles:
            ppl = personlst()
            while roles[priolstloop] != 0:
                if len(ppl[0]) == 0:
                    print(
                        "Not enough people for: "
                        + priolstloop
                        + ", still need "
                        + str(roles[priolstloop])
                        + " more people!"
                    )
                    break

                person = random.choice(ppl[0])
                # Second check to make sure they not already used
                if person not in usedppl:
                    spdsheet.cell(person, 7).value = priolstloop
                    usedppl.append(person)
                    roles[priolstloop] = int(roles[priolstloop]) - 1
                    ppl[0].remove(person)
                else:
                    ppl[0].remove(person)

            roles.pop(priolstloop)
            ss.save("runforrole.xlsx")
    

    # For trailer roles cuase it has prio after decant
    t = "TrailerUL"
    if (t in roles):
        ppl = personlst()
        while roles[t] != 0:

            # Checks if there are any unloaders left
            if (len(ppl[5]) == 0) and (len(ppl[4]) == 0):
                print(
                    "Not enough people for: "
                    + t
                    + ", still need "
                    + str(roles[t])
                    + " more people!"
                )
                break
            
            # Checks for prio unloaders
            if (len(ppl[5]) != 0):
                person = random.choice(ppl[5])  
                if person not in usedppl:
                    spdsheet.cell(person, 7).value = t
                    usedppl.append(person)
                    roles[t] = int(roles[t]) - 1
                    ppl[5].remove(person)
                    spdsheet.cell(person, 6).value = "Yes"
                else: ppl[5].remove(person)
            
            # A group of all the people without roles for the day
            elif (len(ppl[4]) != 0):
                person = random.choice(ppl[4])
                if person not in usedppl:
                    spdsheet.cell(person, 7).value = t
                    usedppl.append(person)
                    roles[t] = int(roles[t]) - 1
                    ppl[4].remove(person)
                    spdsheet.cell(person, 6).value = "Yes"
                else: ppl[4].remove(person)
        
        roles.pop(t)
        ss.save("runforrole.xlsx")

    # For all the non prio roles
    while (len(roles) != 0):
        rles = list(roles.keys())[0]
        regaalst = []
        # Going through the priority
        for row in range(2, spdsheet.max_row + 1):
            rls = spdsheet.cell(row, 4)
            rls1 = re.sub(",", "", str(rls.value))
            if (rles in rls1.split()) and (row not in usedppl): regaalst.append(row)

        while roles[rles] != 0:
            # Checks if anyone is left
            if (len(regaalst) == 0) and (len(ppl[4]) == 0):
                print(
                    "Not enough people for: "
                    + rles
                    + ", still need "
                    + str(roles[rles])
                    + " more people!"
                )
                break 
            # Check for people with prio for the role
            if (len(regaalst) != 0):
                person = random.choice(regaalst)
                if person not in usedppl:
                    spdsheet.cell(person, 7).value = rles
                    usedppl.append(person)
                    roles[rles] = int(roles[rles]) - 1
                    regaalst.remove(person)
                else: regaalst.remove(person)

            # A group of all the people without roles for the day
            elif (len(ppl[4]) != 0):
                person = random.choice(ppl[4])
                if person not in usedppl:
                    spdsheet.cell(person, 7).value = rles
                    usedppl.append(person)
                    roles[rles] = int(roles[rles]) - 1
                    ppl[4].remove(person)
                else: ppl[4].remove(person)
        
        ss.save("runforrole.xlsx")
        roles.pop(rles)

    # For decant cause it has last prio
    ppl = personlst()
    while decant != 0:

        # Checks if there are any decanters left
        if (len(ppl[2]) == 0) and (len(ppl[3]) == 0) and (len(ppl[4]) == 0):
            print(
                "Not enough people for: "
                + d
                + ", still need "
                + str(decant)
                + " more people!"
            )
            break

        # For decanters that have not hit the 5 hours yet
        if (len(ppl[2]) != 0): 
            person = random.choice(ppl[2])  
            if person not in usedppl:
                spdsheet.cell(person, 7).value = d
                usedppl.append(person)
                decant -=1
                ppl[2].remove(person)
                spdsheet.cell(person, 5).value += 10
            else: ppl[2].remove(person)  

        # For decanters that have prio
        elif (len(ppl[3]) != 0):
            person = random.choice(ppl[3])  
            if person not in usedppl:
                spdsheet.cell(person, 7).value = d
                usedppl.append(person)
                decant -=1
                ppl[3].remove(person)
                spdsheet.cell(person, 5).value += 10
            else: ppl[3].remove(person) 

        # A group of all the people without roles for the day
        elif (len(ppl[4]) != 0 ): 
            person = random.choice(ppl[4])
            if person not in usedppl:
                spdsheet.cell(person, 7).value = d
                usedppl.append(person)
                decant -=1
                ppl[4].remove(person)
                spdsheet.cell(person, 5).value += 10
            else: ppl[4].remove(person)
        
    ss.save("runforrole.xlsx")
    
# Start of shift will make sure all the trailer flags are gone and roles are gone
def startofshift():
    for row in range(2, spdsheet.max_row + 1):
        rls = spdsheet.cell(row, 6)
        rls.value = ""
        rls2 = spdsheet.cell(row, 9)
        if str(rls2.value) != "x" and str(rls2.value) != "X":
            rls = spdsheet.cell(row, 7)
            rls.value = ""
    ss.save("runforrole.xlsx")

# Reset Decant will reset decant hours
def resetdecant():
    for row in range(2, spdsheet.max_row + 1):
        rls = spdsheet.cell(row, 5)
        rls.value = 0
    ss.save("runforrole.xlsx")

# Auto tell people unassigned people to go see manager
def unassignedppl():
    for row in range(2, spdsheet.max_row + 1):
        rls = spdsheet.cell(row, 7)
        if (rls is None or rls.value == ""): 
            rls.value = "Unassigned Role, Please see Manager!";
    ss.save("runforrole.xlsx")

#Running Everything
print("Make sure the excel sheet is NOT open else code will not run! Click 'CTRL + C' to end program anytime!")
time.sleep(1)
ss = openpyxl.load_workbook("runforrole.xlsx")
ss1 = ss["Roster Roles"]
spdsheet = ss1
clockedinsheet = openpyxl.load_workbook("tdyatt.xlsx")
spdsheet3 = clockedinsheet.active
pplneededsheet = openpyxl.load_workbook("rolespplneeded.xlsx")
spdsheet2 = pplneededsheet.active

startofshift()
decantrt = input("Do you want to reset decant hours? Enter Y for Yes and N for No! ")
while (True):
    if (decantrt == 'Y' or decantrt == 'N'): break;
    print("Invalid input, try again!") 
    time.sleep(1)
    decantrt = input("Do you want to reset decant hours? Enter Y for Yes and N for No! ")
if (decantrt == 'Y'): resetdecant()
r1 = startup()
retrievepersonb4(r1)
print("DONE FOR START OF SHIFT!")

#Assign all the empty roles 'Unassigned'
unassignedppl()